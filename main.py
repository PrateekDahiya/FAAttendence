import openpyxl
from datetime import datetime
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackContext
from flask import Flask, request
import os
from difflib import get_close_matches
from threading import Thread
import asyncio
import httpx

TOKEN = os.getenv("YOUR_BOT_TOKEN")
WEBHOOK_URL = os.getenv("YOUR_WEBHOOK_URL")
FLASK_PORT = int(os.getenv('FLASK_PORT', 5000))
WEBHOOK_PORT = int(os.getenv('WEBHOOK_PORT', 8443))

app = Flask(__name__)
application = None
student_selection = {}


@app.route(f'/{TOKEN}', methods=['POST'])
def telegram_webhook():
    json_str = request.get_data(as_text=True)
    update = Update.de_json(json_str, application.bot)
    application.update_queue.put_nowait(update)
    return 'ok'


def start_flask():
    app.run(host='0.0.0.0', port=FLASK_PORT)


async def keep_alive():
    while True:
        try:
            async with httpx.AsyncClient() as client:
                await client.get(f"{WEBHOOK_URL}")
            print("Keep-alive ping sent.")
        except Exception as e:
            print(f"Keep-alive failed: {e}")
        await asyncio.sleep(25 * 60)


def format_date_column(date: datetime) -> str:
    return date.strftime("%d%m")


async def update_attendance(roll_number: int, specific_date: str = None, attendance_status: str = 'P') -> str:
    file_path = "FAAtt.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    today_date = datetime.today() if not specific_date else datetime.strptime(specific_date, "%d-%b")
    formatted_date = format_date_column(today_date)

    # Check for existing date column
    date_column = None
    for col in range(5, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value == formatted_date:
            date_column = col
            break

    # If the date column doesn't exist, add a new column
    if date_column is None:
        total_column = sheet.max_column
        if sheet.cell(row=1, column=total_column).value == 'Total':
            total_column -= 1
        sheet.insert_cols(total_column + 1)
        sheet.cell(row=1, column=total_column + 1, value=formatted_date)
        sheet.cell(row=1, column=total_column + 2, value='Total')
        date_column = total_column + 1

    # Find the student row by roll number
    student_row = None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == roll_number:
            student_row = row
            break

    if student_row is None:
        return f"Error: Roll number {roll_number} not found."

    # Update attendance in the found date column
    current_status = sheet.cell(row=student_row, column=date_column).value
    if current_status == attendance_status:
        return f"Already {attendance_status}"

    sheet.cell(row=student_row, column=date_column).value = attendance_status

    # Automatically calculate total attendance using COUNTIF formula
    total_column = sheet.max_column
    count_range = f"E{student_row}:{get_column_letter(total_column - 1)}{student_row}"
    sheet.cell(row=student_row, column=total_column).value = f'=COUNTIF({count_range}, "P")'

    workbook.save(file_path)
    print(f"Marked attendance for roll number {roll_number}.")
    return f"Set {attendance_status} success"


def find_students_by_name(name: str):
    file_path = "FAAtt.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    student_names = [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)]
    matched_names = get_close_matches(name, student_names, n=5, cutoff=0.5)

    matched_students = []
    for row in range(2, sheet.max_row + 1):
        student_name = sheet.cell(row=row, column=1).value
        if student_name in matched_names:
            roll_number = sheet.cell(row=row, column=2).value
            matched_students.append((student_name, roll_number, row))

    return matched_students


async def handle_message(update: Update, context: CallbackContext):
    text = update.message.text.strip()

    if update.message.chat_id in student_selection:
        try:
            selection_index = int(text)
            student_list = student_selection[update.message.chat_id]
            if 0 <= selection_index < len(student_list):
                student_name, roll_number, _ = student_list[selection_index]
                res = await update_attendance(roll_number)
                await update.message.reply_text(res)
                print(f"Marked attendance for {student_name} (Roll: {roll_number}).")
            else:
                await update.message.reply_text("Invalid selection. Try again.")
        except ValueError:
            await update.message.reply_text("Invalid selection. Please enter a number.")
        finally:
            del student_selection[update.message.chat_id]
        return

    try:
        parts = text.split()
        roll_number = None
        specific_date = None
        attendance_status = 'P'

        if len(parts) == 3:
            try:
                specific_date = datetime.strptime(parts[1], "%d-%b").strftime("%d-%b")
                roll_number = int(parts[0])
                if parts[2].lower() in ['a', 'absent']:
                    attendance_status = 'A'
            except ValueError:
                roll_number = None

        elif len(parts) == 2:
            try:
                specific_date = datetime.strptime(parts[1], "%d-%b").strftime("%d-%b")
                roll_number = int(parts[0])
            except ValueError:
                roll_number = int(parts[0]) if parts[0].isdigit() else None
                if parts[1].lower() in ['a', 'absent']:
                    attendance_status = 'A'

        elif len(parts) == 1:
            try:
                roll_number = int(parts[0])
            except ValueError:
                roll_number = None

        if roll_number is not None:
            res = await update_attendance(roll_number, specific_date, attendance_status)
            await update.message.reply_text(res)
        else:
            name = text if len(parts) == 1 else parts[0]
            matched_students = find_students_by_name(name)
            if matched_students:
                if len(matched_students) == 1:
                    student_name, roll_number, _ = matched_students[0]
                    res = await update_attendance(roll_number, specific_date, attendance_status)
                    await update.message.reply_text(res)
                    print(f"Marked attendance for {student_name} (Roll: {roll_number}).")
                else:
                    student_selection[update.message.chat_id] = matched_students
                    student_list = "\n".join(
                        [f"{idx}. {name} (Roll: {roll})" for idx, (name, roll, _) in enumerate(matched_students)])
                    await update.message.reply_text(
                        f"Multiple students found:\n{student_list}\nPlease select a student by number.")
            else:
                await update.message.reply_text(f"No students found matching '{name}'.")
    except ValueError:
        await update.message.reply_text("Invalid input. Please enter a valid roll number or use the /sendfile command.")


async def start(update: Update, context: CallbackContext):
    await update.message.reply_text('Hello! Send me a message, and I will respond. Use /sendfile to get the file.')
    print("Started bot interaction.")


async def send_file(update: Update, context: CallbackContext):
    chat_id = update.message.chat_id
    file_path = './FAAtt.xlsx'
    try:
        with open(file_path, 'rb') as file:
            await context.bot.send_document(chat_id=chat_id, document=file)
            await update.message.reply_text("Here is the file.")
        print(f"Sent attendance file to chat {chat_id}.")
    except Exception as e:
        print(f"Failed to send file: {e}")
        await update.message.reply_text("Failed to send the file.")


async def help_command(update: Update, context: CallbackContext):
    help_text = '''
    Here are the available commands:
    /start - Start the bot
    /sendfile - Get the attendance file
    '''
    await update.message.reply_text(help_text)
    print("Sent help information.")


async def error(update: Update, context: CallbackContext):
    print(f"Error occurred with update {update}")


def main():
    global application

    application = Application.builder().token(TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("sendfile", send_file))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.add_error_handler(error)

    Thread(target=start_flask).start()

    loop = asyncio.get_event_loop()
    loop.create_task(keep_alive())
    application.run_webhook(
        listen="0.0.0.0",
        port=WEBHOOK_PORT,
        webhook_url=f"{WEBHOOK_URL}/{TOKEN}",
    )


if __name__ == '__main__':
    main()
